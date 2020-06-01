import pygame
from shapely.geometry import Point
from shapely.geometry.polygon import Polygon
from openpyxl import load_workbook
import xlsxwriter
pygame.init()


wb2 = load_workbook('hello.xlsx')
ws = wb2.active

WinWidth = 1000
WinHeight = 1000
win = pygame.display.set_mode((WinWidth, WinHeight))
pygame.display.set_caption('pH')

run = True

color = 2
showr = 5
showg = 5
showb = 5
y = 0
cord = [ [ 325946.27048918697983, 273766.753230817557778 ], [ 325972.839882189640775, 273926.270987777446862 ], [ 325980.372286957048345, 273931.420629944710527 ], [ 325991.924007652152795, 273932.615188039490022 ], [ 326008.48215101211099, 273929.925376953615341 ], [ 326022.585010514361784, 273918.421499577525537 ], [ 326035.111265752057079, 273922.362034732941538 ], [ 326129.795054768328555, 273909.426693178596906 ], [ 326231.472820476861671, 273890.075651828548871 ], [ 326246.162918441812508, 273877.879684541665483 ], [ 326276.292659246188123, 273843.315651116252411 ], [ 326286.126762385247275, 273816.275185682985466 ], [ 326286.488272288290318, 273802.664171015552711 ], [ 326254.016610428632703, 273801.32840305409627 ], [ 326248.223163254326209, 273799.561321229906753 ], [ 326255.207047374744434, 273777.98760824103374 ], [ 326254.080925921909511, 273764.103598539077211 ], [ 326242.941075417620596, 273757.987798984569963 ], [ 326175.305306180205662, 273758.660079023917206 ], [ 326175.260305499599781, 273755.30374360823771 ], [ 325946.27048918697983, 273766.753230817557778 ] , [ 326175.260305499599781, 273755.30374360823771 ], [ 326173.528189428208862, 273626.115419429785106 ], [ 325936.857747149013449, 273641.903338780975901 ], [ 325926.050938396947458, 273645.358760554576293 ], [ 325946.27048918697983, 273766.753230817557778 ], [ 326175.260305499599781, 273755.30374360823771 ] ]


# atrod mazākās x, y koordinates
i = 0
minx = cord[1][0]
miny = cord[0][1]
while i < len(cord):
    if minx > cord[i][0]:
        minx = cord[i][0]
    if miny > cord[i][1]:
        miny = cord[i][1]

    i += 1

# koordinates pareikina uz pikseļiem
i = 0
while i < len(cord):
    cord[i][0] -= (minx)
    cord[i][1] -= (miny)
    cord[i][1] = 1000 - cord[i][1] # apgriez otradi

    cord[i][0] += 200
    cord[i][1] -= 200
    i += 1

# atrod visus punktus poligonaa
punkti ={}
i = 0
x=0
while i <= 10000:
    key = i
    if x==100:
        x = 0
        y +=10
    point = Point(x*10, y)
    polygon = Polygon(cord)
    if polygon.contains(point):
        punkti[key] =[[x*10],[y],[0],[0],[0],[0],[0],[0]]
    i += 1
    x +=1

workbook = xlsxwriter.Workbook('hello.xlsx')
worksheet = workbook.add_worksheet()

# ielade punktus no excel
i = 1
while i <= len(punkti)+1:
    for key in punkti:
        c = ws['A'+str(i)]
        if c.value == key:
            pH = ws['E'+str(i)]
            raza = ws['F'+str(i)]
            punkti[key][6][0] = pH.value
            value2 = pH.value
            value2 = ((value2 - 3.5) / 3.3) * 100
            if value2 == 0:
                punkti[key][2][0] = 255
                punkti[key][3][0] = 0
                punkti[key][4][0] = 0
                punkti[key][7][0] = raza.value
            elif 0 < value2 < 50:
                punkti[key][2][0] = 255
                punkti[key][3][0] = int(value2 * 5.1)
                punkti[key][4][0] = 0
                punkti[key][7][0] = raza.value
            elif value2 == 50:
                punkti[key][2][0] = 255
                punkti[key][3][0] = 255
                punkti[key][4][0] = 0
                punkti[key][7][0] = raza.value
            elif 50 < value2 < 100:
                punkti[key][2][0] = int((100 - value2) * 5.1)
                punkti[key][3][0] = 255
                punkti[key][4][0] = 0
                punkti[key][7][0] = raza.value
            elif value2 == 100:
                punkti[key][2][0] = 0
                punkti[key][3][0] = 255
                punkti[key][4][0] = 0
                punkti[key][7][0] = raza.value
            elif 100 < value2 < 150:
                punkti[key][2][0] = 0
                punkti[key][3][0] = 255
                punkti[key][4][0] = int((value2 - 100) * 5.1)
                punkti[key][7][0] = raza.value
            elif value2 == 150:
                punkti[key][2][0] = 0
                punkti[key][3][0] = 255
                punkti[key][4][0] = 255
                punkti[key][7][0] = raza.value
            elif 150 < value2 < 200:
                punkti[key][2][0] = 0
                punkti[key][3][0] = int((200 - value2) * 5.1)
                punkti[key][4][0] = 255
                punkti[key][7][0] = raza.value
            elif value2 == 200:
                punkti[key][2][0] = 0
                punkti[key][3][0] = 0
                punkti[key][4][0] = 255
                punkti[key][7][0] = raza.value
            print(i)
    i += 1

value = 3.5

while run:
    pygame.time.delay(0)

    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            run = False
        move_ticker = 0
        keys = pygame.key.get_pressed()
        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_r:
                color = 2
                print (color)
            if event.key == pygame.K_g:
                color = 3
                print(color)

            if event.key == pygame.K_b:
                color = 4
                print(color)
            if event.key == pygame.K_KP_MINUS:
                value -=0.1
                print(value)
            if event.key == pygame.K_KP_PLUS:
                value += 0.1
                print(value)
            if event.key == pygame.K_KP1:
                showr = 2
                print('RED on')
                print(showr)
            if event.key == pygame.K_KP2:
                showg = 3
                print('GREEN on')
                print(showg)
            if event.key == pygame.K_KP3:
                showb = 4
                print('BLUE on')
                print(showb)
            if event.key == pygame.K_KP4:
                showr = 5
                print('RED off')
                print(showr)
            if event.key == pygame.K_KP5:
                showg =5
                print('GREEN off')
                print(showg)
            if event.key == pygame.K_KP6:
                showb = 5
                print('BLUE off')
                print(showb)

        # atrod id no koordinaatem
        if pygame.mouse.get_pressed()[0]:
            try:
                atlX = event.pos[0]%10
                altY = event.pos[1]%10
                cubeX = event.pos[0] - atlX
                cubeY = event.pos[1] - altY

                cubeY *= 10
                cubeX /= 10
                id = cubeY + cubeX

                #nosaka krasu no vertibas
                if id in punkti:
                    print('in')
                    value2 = ((value - 3.5)/3.3)*100
                    if value2 == 0:
                        punkti[id][2][0] = 255
                        punkti[id][3][0] = 0
                        punkti[id][4][0] = 0
                        punkti[id][6][0] = value
                    elif value2 > 0 and value2 < 50:
                        punkti[id][2][0] = 255
                        punkti[id][3][0] = int(value2 * 5.1)
                        punkti[id][4][0] = 0
                        punkti[id][6][0] = value
                    elif value2 == 50:
                        punkti[id][2][0] = 255
                        punkti[id][3][0] = 255
                        punkti[id][4][0] = 0
                        punkti[id][6][0] = value
                    elif value2 > 50 and value2 < 100:
                        punkti[id][2][0] = int((100-value2)*5.1)
                        punkti[id][3][0] = 255
                        punkti[id][4][0] = 0
                        punkti[id][6][0] = value
                    elif value2 == 100:
                        punkti[id][2][0] = 0
                        punkti[id][3][0] = 255
                        punkti[id][4][0] = 0
                        punkti[id][6][0] = value
                    elif value2 > 100 and value2 < 150:
                        punkti[id][2][0] = 0
                        punkti[id][3][0] = 255
                        punkti[id][4][0] = int((value2-100)*5.1)
                        punkti[id][6][0] = value
                    elif value2 == 150:
                        punkti[id][2][0] = 0
                        punkti[id][3][0] = 255
                        punkti[id][4][0] = 255
                        punkti[id][6][0] = value
                    elif value2 > 150 and value2 < 200:
                        punkti[id][2][0] = 0
                        punkti[id][3][0] = int((200 - value2)*5.1)
                        punkti[id][4][0] = 255
                        punkti[id][6][0] = value
                    elif value2 == 200:
                        punkti[id][2][0] = 0
                        punkti[id][3][0] = 0
                        punkti[id][4][0] = 255
                        punkti[id][6][0] = value

                else:
                    print('Out')

            except AttributeError:
                pass


    # ziimee punktus
    for key in punkti:
        pygame.draw.rect(win, (punkti[key][showr][0], punkti[key][showg][0], punkti[key][showb][0]),
                         (punkti[key][0][0], punkti[key][1][0], 10, 10))

    # lauka kontura
    i = 0
    while i < len(cord)-1:
        pygame.draw.line(win, (255, 255, 255), (cord[i][0], cord[i][1]), (cord[i+1][0], cord[i+1][1]))
        i += 1

    pygame.display.update()


worksheet.write(0,0,'ID')
worksheet.write(0,1,'RED')
worksheet.write(0,2,'GREEN')
worksheet.write(0,3,'BLUE')
worksheet.write(0,4,'pH')
worksheet.write(0,5,'Raža')


# saglabaa
f = 0
for key in punkti:
        print(f)
        print(key)
        worksheet.write(f + 1, 0, key)
        worksheet.write(f + 1, 1, punkti[key][2][0])
        worksheet.write(f + 1, 2, punkti[key][3][0])
        worksheet.write(f + 1, 3, punkti[key][4][0])
        worksheet.write(f + 1, 4, punkti[key][6][0])
        worksheet.write(f + 1, 5, punkti[key][7][0])
        f += 1

workbook.close()

pygame.quit()

